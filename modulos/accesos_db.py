import re
import sqlite3
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

from modulos.base_datos import BaseDatos
from modulos.historial import GestorHistorial


class GestorAccesosDB:
    PREFIJO_IP = "192.168.5."

    CAMPOS_COMPARADOS = {
        "nombre": "Nombre",
        "ip": "IP",
        "mac": "MAC",
        "marca": "Marca",
        "modelo": "Modelo",
        "usuario": "Usuario",
        "ubicacion": "Ubicación"
    }

    def __init__(
        self,
        ruta_db="datos/netinventory.db"
    ):
        self.base_datos = BaseDatos(
            ruta_db
        )

        self.base_datos.crear_tablas()

        self.historial = GestorHistorial(
            self.base_datos
        )

    def normalizar_texto(self, texto):
        if texto is None:
            return ""

        texto = str(texto).strip().lower()

        texto = unicodedata.normalize(
            "NFD",
            texto
        )

        texto = "".join(
            caracter
            for caracter in texto
            if unicodedata.category(caracter) != "Mn"
        )

        return " ".join(
            texto.split()
        )

    def limpiar_valor(self, valor):
        if valor is None:
            return None

        if (
            isinstance(valor, float)
            and valor.is_integer()
        ):
            return int(valor)

        if isinstance(valor, str):
            valor = valor.strip()

            if not valor:
                return None

        return valor

    def validar_ultimo_octeto(self, valor):
        valor = self.limpiar_valor(
            valor
        )

        if valor is None:
            raise ValueError(
                "El último octeto es obligatorio."
            )

        try:
            octeto = int(valor)

        except (
            ValueError,
            TypeError
        ) as error:
            raise ValueError(
                "El último octeto debe ser numérico."
            ) from error

        if not 1 <= octeto <= 254:
            raise ValueError(
                "El último octeto debe estar entre 1 y 254."
            )

        return octeto

    def construir_ip(
        self,
        ultimo_octeto
    ):
        octeto = self.validar_ultimo_octeto(
            ultimo_octeto
        )

        return (
            f"{self.PREFIJO_IP}{octeto}"
        )

    def normalizar_mac(self, mac):
        mac = self.limpiar_valor(
            mac
        )

        if mac is None:
            return None

        caracteres = re.sub(
            r"[^0-9A-Fa-f]",
            "",
            str(mac)
        )

        if len(caracteres) != 12:
            return str(mac).strip()

        return ":".join(
            caracteres[indice:indice + 2]
            for indice in range(
                0,
                12,
                2
            )
        ).upper()

    def limpiar_modelo(
        self,
        marca,
        modelo
    ):
        marca = self.limpiar_valor(
            marca
        )

        modelo = self.limpiar_valor(
            modelo
        )

        if modelo is None:
            return None

        modelo = str(
            modelo
        ).strip()

        if marca is None:
            return modelo

        patron = re.compile(
            rf"^{re.escape(str(marca).strip())}"
            rf"[\s\-_:]*",
            flags=re.IGNORECASE
        )

        modelo_limpio = patron.sub(
            "",
            modelo,
            count=1
        ).strip()

        return modelo_limpio or modelo

    def valores_son_iguales(
        self,
        valor_anterior,
        valor_nuevo
    ):
        """
        Compara valores evitando diferencias irrelevantes
        por espacios o mayúsculas.
        """
        if (
            valor_anterior is None
            and valor_nuevo is None
        ):
            return True

        return (
            str(valor_anterior or "").strip()
            == str(valor_nuevo or "").strip()
        )

    def obtener_cambios(
        self,
        existente,
        datos
    ):
        """
        Devuelve solamente los campos que cambiaron.

        Las contraseñas se registran como modificadas,
        pero sus valores no se guardan.
        """
        cambios = []

        for clave, nombre_visible in (
            self.CAMPOS_COMPARADOS.items()
        ):
            anterior = existente.get(
                clave
            )

            nuevo = datos.get(
                clave
            )

            if not self.valores_son_iguales(
                anterior,
                nuevo
            ):
                cambios.append(
                    {
                        "campo": nombre_visible,
                        "anterior": anterior,
                        "nuevo": nuevo
                    }
                )

        password_nueva = datos.get(
            "password"
        )

        password_anterior = existente.get(
            "password"
        )

        if (
            password_nueva is not None
            and not self.valores_son_iguales(
                password_anterior,
                password_nueva
            )
        ):
            cambios.append(
                {
                    "campo": "Contraseña",
                    "protegido": True
                }
            )

        return cambios

    def listar_todos(self):
        consulta = """
        SELECT *
        FROM accesos_switches
        ORDER BY ultimo_octeto;
        """

        with self.base_datos.conectar() as conexion:
            filas = conexion.execute(
                consulta
            ).fetchall()

        return [
            dict(fila)
            for fila in filas
        ]

    def obtener_por_octeto(
        self,
        ultimo_octeto
    ):
        octeto = self.validar_ultimo_octeto(
            ultimo_octeto
        )

        consulta = """
        SELECT *
        FROM accesos_switches
        WHERE ultimo_octeto = ?;
        """

        with self.base_datos.conectar() as conexion:
            fila = conexion.execute(
                consulta,
                (octeto,)
            ).fetchone()

        if fila is None:
            return None

        return dict(fila)

    def buscar(
        self,
        texto_buscado
    ):
        texto_buscado = self.normalizar_texto(
            texto_buscado
        )

        if not texto_buscado:
            return []

        resultados = []

        for switch in self.listar_todos():
            campos = [
                switch.get("ultimo_octeto"),
                switch.get("ip"),
                switch.get("nombre"),
                switch.get("mac"),
                switch.get("marca"),
                switch.get("modelo"),
                switch.get("usuario"),
                switch.get("ubicacion"),
                switch.get("observaciones")
            ]

            coincide = any(
                texto_buscado
                in self.normalizar_texto(campo)
                for campo in campos
            )

            if coincide:
                resultados.append(
                    switch
                )

        return resultados

    def buscar_por_origen_excel(
        self,
        hoja_excel,
        bloque_excel
    ):
        hoja_normalizada = (
            self.normalizar_texto(
                hoja_excel
            )
        )

        try:
            bloque_excel = int(
                bloque_excel
            )

        except (
            ValueError,
            TypeError
        ):
            return None

        consulta = """
        SELECT *
        FROM accesos_switches
        WHERE bloque_excel = ?;
        """

        with self.base_datos.conectar() as conexion:
            filas = conexion.execute(
                consulta,
                (bloque_excel,)
            ).fetchall()

        for fila in filas:
            switch = dict(fila)

            if (
                self.normalizar_texto(
                    switch.get("hoja_excel")
                )
                == hoja_normalizada
            ):
                return switch

        return None

    def obtener_hoja_por_nombre(
        self,
        libro,
        nombre_buscado
    ):
        nombre_normalizado = (
            self.normalizar_texto(
                nombre_buscado
            )
        )

        for hoja in libro.worksheets:
            if (
                self.normalizar_texto(
                    hoja.title
                )
                == nombre_normalizado
            ):
                return hoja

        return None

    def detectar_columnas_passswitch(
        self,
        hoja
    ):
        equivalencias = {
            "prefijo_ip": {
                "ip",
                "prefijo ip",
                "red"
            },
            "ultimo_octeto": {
                "ultimo octeto",
                "octeto",
                "host",
                "numero",
                "n"
            },
            "nombre": {
                "descripcion",
                "nombre"
            },
            "marca": {
                "marca"
            },
            "modelo": {
                "modelo"
            },
            "ubicacion": {
                "ubicacion"
            },
            "mac": {
                "mac",
                "direccion mac"
            },
            "password": {
                "acceso",
                "password",
                "contraseña",
                "clave"
            }
        }

        equivalencias_normalizadas = {
            campo: {
                self.normalizar_texto(nombre)
                for nombre in nombres
            }
            for campo, nombres
            in equivalencias.items()
        }

        cantidad_filas = (
            hoja.max_row or 1
        )

        cantidad_columnas = (
            hoja.max_column or 1
        )

        ultima_fila_encabezados = min(
            10,
            cantidad_filas
        )

        for numero_fila in range(
            1,
            ultima_fila_encabezados + 1
        ):
            columnas = {}

            for numero_columna in range(
                1,
                cantidad_columnas + 1
            ):
                encabezado = (
                    self.normalizar_texto(
                        hoja.cell(
                            row=numero_fila,
                            column=numero_columna
                        ).value
                    )
                )

                if not encabezado:
                    continue

                for campo, nombres in (
                    equivalencias_normalizadas.items()
                ):
                    if encabezado in nombres:
                        columnas[campo] = (
                            numero_columna
                        )
                        break

            obligatorios = {
                "prefijo_ip",
                "nombre",
                "marca",
                "modelo",
                "ubicacion",
                "mac",
                "password"
            }

            if obligatorios.issubset(
                columnas.keys()
            ):
                if (
                    "ultimo_octeto"
                    not in columnas
                ):
                    columnas[
                        "ultimo_octeto"
                    ] = (
                        columnas["prefijo_ip"]
                        + 1
                    )

                return (
                    numero_fila,
                    columnas
                )

        raise ValueError(
            "No fue posible reconocer los encabezados "
            "de la hoja PASSSWITCH."
        )

    def insertar_desde_passswitch(
        self,
        datos
    ):
        consulta = """
        INSERT INTO accesos_switches (
            ultimo_octeto,
            nombre,
            ip,
            mac,
            marca,
            modelo,
            usuario,
            password,
            ubicacion,
            observaciones
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """

        valores = (
            datos["ultimo_octeto"],
            datos["nombre"],
            datos["ip"],
            datos["mac"],
            datos["marca"],
            datos["modelo"],
            datos["usuario"],
            datos["password"],
            datos["ubicacion"],
            "Importado desde PASSSWITCH"
        )

        try:
            with self.base_datos.conectar() as conexion:
                conexion.execute(
                    consulta,
                    valores
                )

                conexion.commit()

        except sqlite3.IntegrityError as error:
            raise ValueError(
                "Ya existe un switch con esa IP "
                "o último octeto."
            ) from error

        detalle = [
            {
                "campo": "Switch",
                "anterior": None,
                "nuevo": (
                    datos.get("ubicacion")
                    or datos.get("nombre")
                    or datos.get("ip")
                )
            }
        ]

        self.historial.registrar(
            accion="AGREGADO",
            entidad="switch",
            ultimo_octeto=datos.get(
                "ultimo_octeto"
            ),
            ip=datos.get("ip"),
            ubicacion=datos.get(
                "ubicacion"
            ),
            detalle=detalle,
            origen="Importación PASSSWITCH"
        )

    def actualizar_desde_passswitch(
        self,
        datos,
        existente,
        cambios
    ):
        """
        Actualiza un switch conservando su relación.
        """
        consulta = """
        UPDATE accesos_switches
        SET
            nombre = ?,
            ip = ?,
            mac = ?,
            marca = ?,
            modelo = ?,
            usuario = ?,
            password = ?,
            ubicacion = ?,
            fecha_actualizacion = CURRENT_TIMESTAMP
        WHERE ultimo_octeto = ?;
        """

        password_final = (
            datos["password"]
            if datos["password"] is not None
            else existente.get("password")
        )

        valores = (
            datos["nombre"],
            datos["ip"],
            datos["mac"],
            datos["marca"],
            datos["modelo"],
            datos["usuario"],
            password_final,
            datos["ubicacion"],
            datos["ultimo_octeto"]
        )

        with self.base_datos.conectar() as conexion:
            conexion.execute(
                consulta,
                valores
            )

            conexion.commit()

        self.historial.registrar(
            accion="ACTUALIZADO",
            entidad="switch",
            ultimo_octeto=datos.get(
                "ultimo_octeto"
            ),
            ip=datos.get("ip"),
            ubicacion=datos.get(
                "ubicacion"
            ),
            detalle=cambios,
            origen="Importación PASSSWITCH"
        )

    def importar_desde_passswitch(
        self,
        ruta_excel
    ):
        """
        Sincroniza SQLite con PASSSWITCH.
        """
        ruta_excel = Path(
            ruta_excel
        )

        if not ruta_excel.exists():
            raise FileNotFoundError(
                f"No se encontró el Excel: {ruta_excel}"
            )

        libro = load_workbook(
            ruta_excel,
            data_only=True,
            read_only=False
        )

        try:
            hoja = self.obtener_hoja_por_nombre(
                libro,
                "PASSSWITCH"
            )

            if hoja is None:
                raise ValueError(
                    "No se encontró la hoja PASSSWITCH."
                )

            (
                fila_encabezados,
                columnas
            ) = self.detectar_columnas_passswitch(
                hoja
            )

            ultima_fila = (
                hoja.max_row
                or fila_encabezados
            )

            nuevos = 0
            actualizados = 0
            sin_cambios = 0
            ignorados = []
            errores = []

            octetos_en_excel = set()

            for numero_fila in range(
                fila_encabezados + 1,
                ultima_fila + 1
            ):
                valores_fila = {
                    campo: self.limpiar_valor(
                        hoja.cell(
                            row=numero_fila,
                            column=numero_columna
                        ).value
                    )
                    for campo, numero_columna
                    in columnas.items()
                }

                if not any(
                    valores_fila.values()
                ):
                    continue

                try:
                    octeto = (
                        self.validar_ultimo_octeto(
                            valores_fila.get(
                                "ultimo_octeto"
                            )
                        )
                    )

                except ValueError as error:
                    ignorados.append(
                        {
                            "fila": numero_fila,
                            "motivo": str(error)
                        }
                    )

                    continue

                octetos_en_excel.add(
                    octeto
                )

                nombre = (
                    valores_fila.get("nombre")
                    or valores_fila.get(
                        "ubicacion"
                    )
                    or f"Switch {octeto}"
                )

                marca = valores_fila.get(
                    "marca"
                )

                datos = {
                    "ultimo_octeto": octeto,
                    "ip": self.construir_ip(
                        octeto
                    ),
                    "nombre": nombre,
                    "mac": self.normalizar_mac(
                        valores_fila.get("mac")
                    ),
                    "marca": marca,
                    "modelo": self.limpiar_modelo(
                        marca,
                        valores_fila.get(
                            "modelo"
                        )
                    ),
                    "usuario": "admin",
                    "ubicacion": valores_fila.get(
                        "ubicacion"
                    ),
                    "password": valores_fila.get(
                        "password"
                    )
                }

                existente = self.obtener_por_octeto(
                    octeto
                )

                try:
                    if existente is None:
                        self.insertar_desde_passswitch(
                            datos
                        )

                        nuevos += 1
                        continue

                    cambios = self.obtener_cambios(
                        existente,
                        datos
                    )

                    if not cambios:
                        sin_cambios += 1
                        continue

                    self.actualizar_desde_passswitch(
                        datos=datos,
                        existente=existente,
                        cambios=cambios
                    )

                    actualizados += 1

                except Exception as error:
                    errores.append(
                        {
                            "fila": numero_fila,
                            "octeto": octeto,
                            "motivo": str(error)
                        }
                    )

            switches_ausentes = [
                switch
                for switch in self.listar_todos()
                if switch.get(
                    "ultimo_octeto"
                ) not in octetos_en_excel
            ]

            return {
                "nuevos": nuevos,
                "actualizados": actualizados,
                "sin_cambios": sin_cambios,
                "ignorados": ignorados,
                "errores": errores,
                "ausentes": switches_ausentes
            }

        finally:
            libro.close()

    def eliminar_switches_ausentes(
        self,
        switches_ausentes
    ):
        """
        Elimina switches confirmados como ausentes
        y registra la acción en el historial.
        """
        eliminados = 0
        errores = []

        for switch in switches_ausentes:
            octeto = switch.get(
                "ultimo_octeto"
            )

            if octeto is None:
                errores.append(
                    {
                        "ip": switch.get("ip"),
                        "motivo": (
                            "El switch no tiene "
                            "último octeto."
                        )
                    }
                )

                continue

            try:
                with self.base_datos.conectar() as conexion:
                    cursor = conexion.execute(
                        """
                        DELETE FROM accesos_switches
                        WHERE ultimo_octeto = ?;
                        """,
                        (octeto,)
                    )

                    conexion.commit()

                if cursor.rowcount > 0:
                    eliminados += 1

                    self.historial.registrar(
                        accion="ELIMINADO",
                        entidad="switch",
                        ultimo_octeto=octeto,
                        ip=switch.get("ip"),
                        ubicacion=switch.get(
                            "ubicacion"
                        ),
                        detalle=[
                            {
                                "campo": "Switch",
                                "anterior": (
                                    switch.get(
                                        "ubicacion"
                                    )
                                    or switch.get(
                                        "nombre"
                                    )
                                    or switch.get(
                                        "ip"
                                    )
                                ),
                                "nuevo": None
                            }
                        ],
                        origen=(
                            "Sincronización PASSSWITCH"
                        )
                    )

            except sqlite3.Error as error:
                errores.append(
                    {
                        "ip": switch.get("ip"),
                        "motivo": str(error)
                    }
                )

        return {
            "detectados": len(
                switches_ausentes
            ),
            "eliminados": eliminados,
            "errores": errores
        }

    def valor_visible(
        self,
        valor
    ):
        if valor is None:
            return "Sin información"

        return valor

    def mostrar_resultados(
        self,
        resultados
    ):
        if not resultados:
            print(
                "\nNo se encontraron switches."
            )
            return

        print(
            f"\nSwitches encontrados: "
            f"{len(resultados)}"
        )

        for numero, switch in enumerate(
            resultados,
            start=1
        ):
            modelo = self.limpiar_modelo(
                switch.get("marca"),
                switch.get("modelo")
            )

            print(
                "\n========================================"
            )
            print(
                f"Resultado {numero}"
            )
            print(
                "IP: "
                f"{self.valor_visible(switch.get('ip'))}"
            )
            print(
                "Ubicación: "
                f"{self.valor_visible(switch.get('ubicacion'))}"
            )
            print(
                "MAC: "
                f"{self.valor_visible(switch.get('mac'))}"
            )
            print(
                "Marca: "
                f"{self.valor_visible(switch.get('marca'))}"
            )
            print(
                "Modelo: "
                f"{self.valor_visible(modelo)}"
            )
            print(
                "Usuario: "
                f"{self.valor_visible(switch.get('usuario'))}"
            )
            print(
                "Contraseña: "
                f"{self.valor_visible(switch.get('password'))}"
            )

            observaciones = switch.get(
                "observaciones"
            )

            if observaciones:
                print(
                    f"Observaciones: {observaciones}"
                )