import unicodedata
from collections import Counter

from openpyxl import load_workbook


class Inventario:
    def __init__(self, ruta_excel):
        self.ruta_excel = ruta_excel
        self.libro = None

        self.tablas_detectadas = []
        self.registros = []

        self.filas_incompletas = []
        self.duplicados_detectados = []

    def limpiar_valor(self, valor):
        """
        Limpia un valor obtenido desde Excel.
        """
        if valor is None:
            return None

        if isinstance(valor, float) and valor.is_integer():
            return int(valor)

        if isinstance(valor, str):
            valor = valor.strip()

            if valor == "":
                return None

        return valor

    def normalizar_texto(self, texto):
        """
        Normaliza textos para compararlos ignorando:
        - mayúsculas y minúsculas;
        - tildes;
        - espacios sobrantes.
        """
        if texto is None:
            return ""

        texto = str(texto).strip().lower()
        texto = unicodedata.normalize("NFD", texto)

        texto = "".join(
            caracter
            for caracter in texto
            if unicodedata.category(caracter) != "Mn"
        )

        return " ".join(texto.split())

    def normalizar_tipo(self, tipo):
        """
        Convierte distintas formas de escribir un tipo
        en una categoría común.
        """
        texto = self.normalizar_texto(tipo)

        if not texto:
            return "Sin tipo"

        equivalencias = {
            "antena": "Antena",
            "ap": "Antena",
            "access point": "Antena",

            "camara": "Cámara",
            "camaras": "Cámara",

            "telefono": "Teléfono IP",
            "telefono ip": "Teléfono IP",

            "troncal": "Troncal",
            "trunk": "Troncal",

            "equipo": "Equipo",
            "equipo alumnos": "Equipo"
        }

        return equivalencias.get(
            texto,
            str(tipo).strip()
        )

    def cargar_excel(self):
        """
        Abre el archivo Excel principal.
        """
        try:
            self.libro = load_workbook(
                self.ruta_excel,
                data_only=True
            )

            print("Excel abierto correctamente.")
            print(
                f"Hojas encontradas: "
                f"{len(self.libro.sheetnames)}"
            )

        except FileNotFoundError as error:
            raise FileNotFoundError(
                f"No se encontró el archivo: "
                f"{self.ruta_excel}"
            ) from error

        except PermissionError as error:
            raise PermissionError(
                "No se pudo abrir el Excel. "
                "Comprueba que no esté abierto en Excel."
            ) from error

    def es_encabezado_equipo(self, valor):
        """
        Indica si una celda parece ser el encabezado
        de la columna que contiene el nombre del equipo.
        """
        texto = self.normalizar_texto(valor)

        return texto in {
            "equipo",
            "equipo sala",
            "descripcion",
            "dispositivo"
        }

    def es_encabezado_puerto(self, valor):
        """
        Indica si una celda parece ser el encabezado
        del puerto del switch.
        """
        texto = self.normalizar_texto(valor)

        return texto in {
            "boca",
            "boca aruba",
            "boca cisco",
            "puerto",
            "puerto switch",
            "puerto aruba",
            "puerto cisco"
        }

    def buscar_columna_tipo(
        self,
        hoja,
        fila_encabezado,
        columna_equipo
    ):
        """
        Busca una columna de tipo situada antes
        de la columna del equipo.
        """
        for columna in range(
            columna_equipo - 1,
            max(0, columna_equipo - 3),
            -1
        ):
            if columna < 1:
                continue

            valor = hoja.cell(
                row=fila_encabezado,
                column=columna
            ).value

            if self.normalizar_texto(valor) in {
                "patch panel",
                "tipo",
                "categoria",
                "clase"
            }:
                return columna

        return None

    def buscar_columna_patch(
        self,
        hoja,
        fila_encabezado,
        columna_equipo,
        columna_puerto
    ):
        """
        Busca la columna de boca del patch panel.
        """
        for columna in range(
            columna_equipo + 1,
            columna_puerto
        ):
            valor = hoja.cell(
                row=fila_encabezado,
                column=columna
            ).value

            if self.normalizar_texto(valor) in {
                "boca patch",
                "puerto patch",
                "patch",
                "n patch"
            }:
                return columna

        return None

    def buscar_columna_vlan(
        self,
        hoja,
        fila_encabezado,
        columna_puerto
    ):
        """
        Busca una columna VLAN cercana al puerto.
        """
        columna_final = min(
            columna_puerto + 3,
            hoja.max_column
        )

        for columna in range(
            columna_puerto + 1,
            columna_final + 1
        ):
            valor = hoja.cell(
                row=fila_encabezado,
                column=columna
            ).value

            if self.normalizar_texto(valor) == "vlan":
                return columna

        return None

    def tabla_ya_detectada(self, nueva_tabla):
        """
        Evita guardar dos veces la misma tabla.
        """
        for tabla in self.tablas_detectadas:
            misma_tabla = (
                tabla["hoja"] == nueva_tabla["hoja"]
                and tabla["fila_encabezado"]
                == nueva_tabla["fila_encabezado"]
                and tabla["columna_equipo"]
                == nueva_tabla["columna_equipo"]
                and tabla["columna_puerto"]
                == nueva_tabla["columna_puerto"]
            )

            if misma_tabla:
                return True

        return False

    def ultima_fila_con_datos(self, hoja):
        """
        Busca la última fila que contiene información real.
        """
        for numero_fila in range(
            hoja.max_row,
            0,
            -1
        ):
            tiene_datos = any(
                hoja.cell(
                    row=numero_fila,
                    column=columna
                ).value is not None
                for columna in range(
                    1,
                    hoja.max_column + 1
                )
            )

            if tiene_datos:
                return numero_fila

        return 1

    def detectar_tablas_switches(self):
        """
        Recorre el Excel y detecta automáticamente
        los bloques que parecen representar switches.
        """
        if self.libro is None:
            raise RuntimeError(
                "Primero debes ejecutar cargar_excel()."
            )

        self.tablas_detectadas = []

        hojas_excluidas = {
            "accesos a switch",
            "ap",
            "all sw",
            "passswitch"
        }

        for hoja in self.libro.worksheets:
            nombre_normalizado = self.normalizar_texto(
                hoja.title
            )

            if nombre_normalizado in hojas_excluidas:
                continue

            ultima_fila = self.ultima_fila_con_datos(
                hoja
            )

            for numero_fila in range(
                1,
                ultima_fila + 1
            ):
                columnas_equipo = []
                columnas_puerto = []

                for numero_columna in range(
                    1,
                    hoja.max_column + 1
                ):
                    valor = hoja.cell(
                        row=numero_fila,
                        column=numero_columna
                    ).value

                    if self.es_encabezado_equipo(valor):
                        columnas_equipo.append(
                            numero_columna
                        )

                    if self.es_encabezado_puerto(valor):
                        columnas_puerto.append(
                            numero_columna
                        )

                for columna_puerto in columnas_puerto:
                    equipos_anteriores = [
                        columna
                        for columna in columnas_equipo
                        if columna < columna_puerto
                    ]

                    if not equipos_anteriores:
                        continue

                    columna_equipo = max(
                        equipos_anteriores
                    )

                    tabla = {
                        "hoja": hoja.title.strip(),
                        "fila_encabezado": numero_fila,
                        "fila_inicio": numero_fila + 1,
                        "fila_fin": ultima_fila,

                        "columna_tipo": (
                            self.buscar_columna_tipo(
                                hoja,
                                numero_fila,
                                columna_equipo
                            )
                        ),

                        "columna_equipo": columna_equipo,

                        "columna_patch": (
                            self.buscar_columna_patch(
                                hoja,
                                numero_fila,
                                columna_equipo,
                                columna_puerto
                            )
                        ),

                        "columna_puerto": columna_puerto,

                        "columna_vlan": (
                            self.buscar_columna_vlan(
                                hoja,
                                numero_fila,
                                columna_puerto
                            )
                        )
                    }

                    if not self.tabla_ya_detectada(tabla):
                        self.tablas_detectadas.append(
                            tabla
                        )

        self.calcular_finales_tablas()

        print(
            "Tablas de switches detectadas: "
            f"{len(self.tablas_detectadas)}"
        )

        return self.tablas_detectadas

    def calcular_finales_tablas(self):
        """
        Si existen dos tablas verticales usando las mismas
        columnas, hace que la primera termine antes de la segunda.
        """
        for tabla_actual in self.tablas_detectadas:
            siguientes = [
                tabla
                for tabla in self.tablas_detectadas
                if (
                    tabla["hoja"]
                    == tabla_actual["hoja"]
                    and tabla["fila_encabezado"]
                    > tabla_actual["fila_encabezado"]
                    and tabla["columna_puerto"]
                    == tabla_actual["columna_puerto"]
                )
            ]

            if siguientes:
                siguiente = min(
                    siguientes,
                    key=lambda tabla: tabla[
                        "fila_encabezado"
                    ]
                )

                tabla_actual["fila_fin"] = (
                    siguiente["fila_encabezado"] - 1
                )

    def obtener_hoja(self, nombre_buscado):
        """
        Obtiene una hoja ignorando mayúsculas,
        espacios y tildes.
        """
        nombre_normalizado = self.normalizar_texto(
            nombre_buscado
        )

        for hoja in self.libro.worksheets:
            if (
                self.normalizar_texto(hoja.title)
                == nombre_normalizado
            ):
                return hoja

        return None

    def obtener_valor_celda(
        self,
        hoja,
        fila,
        columna
    ):
        """
        Obtiene y limpia una celda.
        """
        if columna is None:
            return None

        valor = hoja.cell(
            row=fila,
            column=columna
        ).value

        return self.limpiar_valor(valor)

    def registro_tiene_informacion(
        self,
        tipo,
        equipo,
        boca_patch,
        vlan
    ):
        """
        Comprueba que exista algún dato asociado al puerto.
        """
        return any(
            valor is not None
            for valor in [
                tipo,
                equipo,
                boca_patch,
                vlan
            ]
        )

    def cargar_registros_switches(self):
        """
        Lee los registros de todas las tablas detectadas.
        """
        if self.libro is None:
            raise RuntimeError(
                "Primero debes ejecutar cargar_excel()."
            )

        if not self.tablas_detectadas:
            raise RuntimeError(
                "Primero debes ejecutar "
                "detectar_tablas_switches()."
            )

        self.registros = []
        self.filas_incompletas = []
        self.duplicados_detectados = []

        contador_bloques = {}
        claves_registradas = {}

        tablas_ordenadas = sorted(
            self.tablas_detectadas,
            key=lambda tabla: (
                tabla["hoja"],
                tabla["fila_encabezado"],
                tabla["columna_puerto"]
            )
        )

        for tabla in tablas_ordenadas:
            nombre_hoja = tabla["hoja"]

            contador_bloques[nombre_hoja] = (
                contador_bloques.get(
                    nombre_hoja,
                    0
                )
                + 1
            )

            numero_bloque = contador_bloques[
                nombre_hoja
            ]

            hoja = self.obtener_hoja(
                nombre_hoja
            )

            if hoja is None:
                continue

            for numero_fila in range(
                tabla["fila_inicio"],
                tabla["fila_fin"] + 1
            ):
                puerto_switch = self.obtener_valor_celda(
                    hoja,
                    numero_fila,
                    tabla["columna_puerto"]
                )

                if not isinstance(
                    puerto_switch,
                    int
                ):
                    continue

                tipo_original = self.obtener_valor_celda(
                    hoja,
                    numero_fila,
                    tabla["columna_tipo"]
                )

                equipo = self.obtener_valor_celda(
                    hoja,
                    numero_fila,
                    tabla["columna_equipo"]
                )

                boca_patch = self.obtener_valor_celda(
                    hoja,
                    numero_fila,
                    tabla["columna_patch"]
                )

                vlan = self.obtener_valor_celda(
                    hoja,
                    numero_fila,
                    tabla["columna_vlan"]
                )

                if not self.registro_tiene_informacion(
                    tipo_original,
                    equipo,
                    boca_patch,
                    vlan
                ):
                    self.filas_incompletas.append(
                        {
                            "hoja": nombre_hoja,
                            "bloque": numero_bloque,
                            "fila_excel": numero_fila,
                            "puerto_switch": puerto_switch
                        }
                    )

                    continue

                clave = (
                    nombre_hoja,
                    numero_bloque,
                    puerto_switch
                )

                if clave in claves_registradas:
                    registro_anterior = (
                        claves_registradas[clave]
                    )

                    self.duplicados_detectados.append(
                        {
                            "hoja": nombre_hoja,
                            "bloque": numero_bloque,
                            "puerto_switch": puerto_switch,
                            "primera_fila": (
                                registro_anterior[
                                    "fila_excel"
                                ]
                            ),
                            "segunda_fila": numero_fila,
                            "primer_equipo": (
                                registro_anterior[
                                    "equipo"
                                ]
                            ),
                            "segundo_equipo": equipo
                        }
                    )

                    continue

                registro = {
                    "hoja": nombre_hoja,
                    "bloque": numero_bloque,
                    "fila_excel": numero_fila,

                    "tipo_original": tipo_original,
                    "tipo": self.normalizar_tipo(
                        tipo_original
                    ),

                    "equipo": equipo,
                    "boca_patch": boca_patch,
                    "puerto_switch": puerto_switch,
                    "vlan": vlan
                }

                claves_registradas[clave] = registro
                self.registros.append(registro)

        print(
            "Registros de puertos cargados: "
            f"{len(self.registros)}"
        )

        print(
            "Filas incompletas ignoradas: "
            f"{len(self.filas_incompletas)}"
        )

        print(
            "Puertos duplicados detectados: "
            f"{len(self.duplicados_detectados)}"
        )

        return self.registros

    def equipo_esta_disponible(self, equipo):
        """
        Comprueba si un registro contiene la palabra disponible.
        """
        return (
            "disponible"
            in self.normalizar_texto(equipo)
        )

    def equipo_es_desconocido(self, equipo):
        """
        Comprueba si un equipo está marcado como desconocido.
        """
        return self.normalizar_texto(equipo) in {
            "??",
            "?",
            "desconocido",
            "sin identificar",
            "no identificado"
        }

    def mostrar_resumen_general(self):
        """
        Muestra las estadísticas generales del inventario.
        """
        if not self.registros:
            print("\nNo hay registros cargados.")
            return

        total = len(self.registros)

        con_equipo = sum(
            registro["equipo"] is not None
            for registro in self.registros
        )

        sin_equipo = total - con_equipo

        disponibles = sum(
            self.equipo_esta_disponible(
                registro["equipo"]
            )
            for registro in self.registros
        )

        desconocidos = sum(
            self.equipo_es_desconocido(
                registro["equipo"]
            )
            for registro in self.registros
        )

        con_vlan = sum(
            registro["vlan"] is not None
            for registro in self.registros
        )

        hojas = {
            registro["hoja"]
            for registro in self.registros
        }

        bloques = {
            (
                registro["hoja"],
                registro["bloque"]
            )
            for registro in self.registros
        }

        tipos = Counter(
            registro["tipo"]
            for registro in self.registros
        )

        registros_por_hoja = Counter(
            registro["hoja"]
            for registro in self.registros
        )

        print("\n========== RESUMEN GENERAL ==========")
        print(
            f"Hojas con puertos cargados: "
            f"{len(hojas)}"
        )
        print(
            f"Bloques cargados: {len(bloques)}"
        )
        print(
            f"Registros documentados: {total}"
        )
        print(
            f"Registros con equipo: {con_equipo}"
        )
        print(
            f"Registros sin equipo: {sin_equipo}"
        )
        print(
            f"Marcados como disponibles: {disponibles}"
        )
        print(
            f"Marcados como desconocidos: {desconocidos}"
        )
        print(
            f"Registros con VLAN: {con_vlan}"
        )
        print(
            f"Registros sin VLAN: {total - con_vlan}"
        )

        print(
            "\n========== REGISTROS POR TIPO =========="
        )

        for tipo, cantidad in sorted(
            tipos.items()
        ):
            print(
                f"- {tipo}: {cantidad}"
            )

        print(
            "\n========== REGISTROS POR HOJA =========="
        )

        for hoja, cantidad in sorted(
            registros_por_hoja.items()
        ):
            print(
                f"- {hoja}: {cantidad}"
            )

    def mostrar_advertencias(self):
        """
        Muestra todos los registros que necesitan revisión.
        Siempre imprime un resultado.
        """
        print(
            "\n========== ADVERTENCIAS DEL INVENTARIO =========="
        )

        cantidad_incompletas = len(
            self.filas_incompletas
        )

        cantidad_repetidos = len(
            self.duplicados_detectados
        )

        registros_sin_equipo = [
            registro
            for registro in self.registros
            if registro.get("equipo") is None
        ]

        registros_desconocidos = [
            registro
            for registro in self.registros
            if self.equipo_es_desconocido(
                registro.get("equipo")
            )
        ]

        print(
            "Filas incompletas ignoradas: "
            f"{cantidad_incompletas}"
        )

        print(
            "Puertos repetidos para revisión: "
            f"{cantidad_repetidos}"
        )

        print(
            "Registros sin equipo identificado: "
            f"{len(registros_sin_equipo)}"
        )

        print(
            "Registros marcados como desconocidos: "
            f"{len(registros_desconocidos)}"
        )

        if (
            cantidad_incompletas == 0
            and cantidad_repetidos == 0
            and not registros_sin_equipo
            and not registros_desconocidos
        ):
            print(
                "\nNo se encontraron advertencias "
                "en el inventario."
            )
            return

        if self.filas_incompletas:
            print(
                "\n---------- FILAS INCOMPLETAS ----------"
            )

            for registro in self.filas_incompletas:
                print(
                    f"Hoja: {registro.get('hoja')} | "
                    f"Bloque: {registro.get('bloque')} | "
                    f"Fila Excel: "
                    f"{registro.get('fila_excel')} | "
                    f"Puerto: "
                    f"{registro.get('puerto_switch')}"
                )

        if self.duplicados_detectados:
            print(
                "\n---------- PUERTOS REPETIDOS ----------"
            )

            for duplicado in self.duplicados_detectados:
                print(
                    f"Hoja: {duplicado.get('hoja')} | "
                    f"Bloque: {duplicado.get('bloque')} | "
                    f"Puerto: "
                    f"{duplicado.get('puerto_switch')} | "
                    f"Filas: "
                    f"{duplicado.get('primera_fila')} y "
                    f"{duplicado.get('segunda_fila')} | "
                    f"Equipos: "
                    f"{duplicado.get('primer_equipo')} / "
                    f"{duplicado.get('segundo_equipo')}"
                )

        if registros_sin_equipo:
            print(
                "\n---------- REGISTROS SIN EQUIPO ----------"
            )

            for registro in registros_sin_equipo:
                print(
                    f"Hoja: {registro.get('hoja')} | "
                    f"Bloque: {registro.get('bloque')} | "
                    f"Puerto: "
                    f"{registro.get('puerto_switch')} | "
                    f"Fila Excel: "
                    f"{registro.get('fila_excel')}"
                )

        if registros_desconocidos:
            print(
                "\n---------- EQUIPOS DESCONOCIDOS ----------"
            )

            for registro in registros_desconocidos:
                print(
                    f"Hoja: {registro.get('hoja')} | "
                    f"Bloque: {registro.get('bloque')} | "
                    f"Puerto: "
                    f"{registro.get('puerto_switch')} | "
                    f"Equipo: "
                    f"{registro.get('equipo')} | "
                    f"Fila Excel: "
                    f"{registro.get('fila_excel')}"
                )

        print(
            "\nFin de las advertencias."
        )
    