from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import (
    BarChart,
    DoughnutChart,
    PieChart,
    Reference
)
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side
)
from openpyxl.utils import get_column_letter

from config import (
    AREA_RESPONSABLE,
    COLOR_ADVERTENCIA,
    COLOR_BORDE,
    COLOR_CORRECTO,
    COLOR_CRITICO,
    COLOR_ENCABEZADO,
    COLOR_INFORMATIVO,
    COLOR_TEXTO_CLARO,
    COLOR_TITULO,
    NOMBRE_APLICACION,
    ORGANIZACION,
    VERSION_APLICACION
)
from modulos.revision_incompletos import (
    RevisorIncompletos
)


class ExportadorExcel:
    """
    Genera un reporte administrativo y técnico de la red.

    El reporte contiene indicadores, análisis, gráficos,
    validaciones, switches y datos detallados.

    Nunca modifica el archivo Excel original.
    """

    def __init__(
        self,
        carpeta_reportes="reportes",
        nombre_archivo_origen=None,
        analizador_reportes=None,
        validador_inventario=None
    ):
        self.carpeta_reportes = Path(
            carpeta_reportes
        )

        self.nombre_archivo_origen = (
            nombre_archivo_origen
        )

        self.analizador_reportes = (
            analizador_reportes
        )

        self.validador_inventario = (
            validador_inventario
        )

        self.carpeta_reportes.mkdir(
            parents=True,
            exist_ok=True
        )

        self.borde_fino = Border(
            left=Side(
                style="thin",
                color=COLOR_BORDE
            ),
            right=Side(
                style="thin",
                color=COLOR_BORDE
            ),
            top=Side(
                style="thin",
                color=COLOR_BORDE
            ),
            bottom=Side(
                style="thin",
                color=COLOR_BORDE
            )
        )

    @staticmethod
    def valor_visible(valor):
        """
        Convierte valores None en texto vacío.
        """
        if valor is None:
            return ""

        return valor

    def aplicar_titulo(
        self,
        hoja,
        titulo,
        cantidad_columnas
    ):
        """
        Agrega un título superior a una hoja.
        """
        hoja.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=cantidad_columnas
        )

        celda = hoja.cell(
            row=1,
            column=1,
            value=titulo
        )

        celda.font = Font(
            bold=True,
            color=COLOR_TEXTO_CLARO,
            size=15
        )

        celda.fill = PatternFill(
            fill_type="solid",
            fgColor=COLOR_TITULO
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        hoja.row_dimensions[1].height = 28

    def aplicar_encabezados(
        self,
        hoja,
        fila,
        encabezados
    ):
        """
        Escribe y formatea encabezados.
        """
        for columna, encabezado in enumerate(
            encabezados,
            start=1
        ):
            celda = hoja.cell(
                row=fila,
                column=columna,
                value=encabezado
            )

            celda.font = Font(
                bold=True,
                color=COLOR_TEXTO_CLARO
            )

            celda.fill = PatternFill(
                fill_type="solid",
                fgColor=COLOR_ENCABEZADO
            )

            celda.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            celda.border = self.borde_fino

    def aplicar_formato_tabla(
        self,
        hoja,
        fila_inicio,
        fila_fin,
        columna_fin
    ):
        """
        Aplica bordes y ajuste de texto a una tabla.
        """
        for fila in hoja.iter_rows(
            min_row=fila_inicio,
            max_row=fila_fin,
            min_col=1,
            max_col=columna_fin
        ):
            for celda in fila:
                celda.border = self.borde_fino

                celda.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

    @staticmethod
    def ajustar_columnas(
        hoja,
        ancho_maximo=42
    ):
        """
        Ajusta el ancho de las columnas.
        """
        for columna in range(
            1,
            hoja.max_column + 1
        ):
            longitud_maxima = 0

            for fila in range(
                1,
                hoja.max_row + 1
            ):
                valor = hoja.cell(
                    row=fila,
                    column=columna
                ).value

                if valor is None:
                    continue

                longitud_maxima = max(
                    longitud_maxima,
                    len(str(valor))
                )

            ancho = min(
                longitud_maxima + 3,
                ancho_maximo
            )

            hoja.column_dimensions[
                get_column_letter(
                    columna
                )
            ].width = max(
                ancho,
                12
            )

    @staticmethod
    def congelar_y_filtrar(
        hoja,
        fila_encabezados,
        ultima_fila,
        ultima_columna
    ):
        """
        Congela los encabezados y habilita filtros.
        """
        hoja.freeze_panes = (
            f"A{fila_encabezados + 1}"
        )

        if ultima_fila < fila_encabezados:
            return

        letra_final = get_column_letter(
            ultima_columna
        )

        hoja.auto_filter.ref = (
            f"A{fila_encabezados}:"
            f"{letra_final}{ultima_fila}"
        )

    @staticmethod
    def aplicar_relleno(
        celda,
        color
    ):
        """
        Aplica un color de fondo a una celda.
        """
        celda.fill = PatternFill(
            fill_type="solid",
            fgColor=color
        )

    def aplicar_color_fila(
        self,
        hoja,
        fila,
        columna_final,
        color
    ):
        """
        Aplica color a toda una fila.
        """
        for columna in range(
            1,
            columna_final + 1
        ):
            self.aplicar_relleno(
                hoja.cell(
                    row=fila,
                    column=columna
                ),
                color
            )

    def crear_hoja_resumen_ejecutivo(
        self,
        libro
    ):
        """
        Presenta indicadores, hallazgos y recomendaciones.
        """
        hoja = libro.active
        hoja.title = "Resumen ejecutivo"

        self.aplicar_titulo(
            hoja,
            "RESUMEN EJECUTIVO DE LA RED",
            6
        )

        indicadores = (
            self.analizador_reportes
            .obtener_indicadores_generales()
        )

        datos_identidad = [
            (
                "Aplicación",
                NOMBRE_APLICACION
            ),
            (
                "Versión",
                VERSION_APLICACION
            ),
            (
                "Organización",
                ORGANIZACION
            ),
            (
                "Área responsable",
                AREA_RESPONSABLE
            ),
            (
                "Fecha de generación",
                datetime.now().strftime(
                    "%d-%m-%Y %H:%M:%S"
                )
            ),
            (
                "Archivo de origen",
                self.nombre_archivo_origen
                or "Sin información"
            )
        ]

        fila = 3

        for nombre, resultado in datos_identidad:
            hoja.cell(
                row=fila,
                column=1,
                value=nombre
            )

            hoja.cell(
                row=fila,
                column=2,
                value=resultado
            )

            fila += 1

        hoja.merge_cells(
            start_row=10,
            start_column=1,
            end_row=10,
            end_column=6
        )

        celda_indicadores = hoja.cell(
            row=10,
            column=1,
            value="INDICADORES PRINCIPALES"
        )

        celda_indicadores.font = Font(
            bold=True,
            color=COLOR_TEXTO_CLARO
        )

        self.aplicar_relleno(
            celda_indicadores,
            COLOR_TITULO
        )

        celda_indicadores.alignment = Alignment(
            horizontal="center"
        )

        indicadores_visibles = [
            (
                "Puertos documentados",
                indicadores["total_registros"]
            ),
            (
                "Puertos ocupados",
                indicadores["puertos_ocupados"]
            ),
            (
                "Puertos disponibles",
                indicadores["puertos_disponibles"]
            ),
            (
                "Puertos sin definir",
                indicadores["puertos_sin_definir"]
            ),
            (
                "Ocupación estimada",
                indicadores[
                    "porcentaje_ocupacion"
                ] / 100
            ),
            (
                "Disponibilidad",
                indicadores[
                    "porcentaje_disponibilidad"
                ] / 100
            ),
            (
                "Cobertura de VLAN",
                indicadores[
                    "porcentaje_con_vlan"
                ] / 100
            ),
            (
                "Incidencias críticas",
                indicadores[
                    "validaciones_criticas"
                ]
            ),
            (
                "Advertencias",
                indicadores["advertencias"]
            ),
            (
                "Switches sin relación",
                indicadores[
                    "switches_sin_relacion"
                ]
            )
        ]

        posiciones = [
            (12, 1),
            (12, 3),
            (12, 5),
            (15, 1),
            (15, 3),
            (15, 5),
            (18, 1),
            (18, 3),
            (18, 5),
            (21, 1)
        ]

        for (
            etiqueta,
            valor
        ), (
            fila,
            columna
        ) in zip(
            indicadores_visibles,
            posiciones
        ):
            hoja.merge_cells(
                start_row=fila,
                start_column=columna,
                end_row=fila,
                end_column=columna + 1
            )

            hoja.merge_cells(
                start_row=fila + 1,
                start_column=columna,
                end_row=fila + 2,
                end_column=columna + 1
            )

            celda_titulo = hoja.cell(
                row=fila,
                column=columna,
                value=etiqueta
            )

            celda_valor = hoja.cell(
                row=fila + 1,
                column=columna,
                value=valor
            )

            celda_titulo.font = Font(
                bold=True,
                color=COLOR_TEXTO_CLARO
            )

            self.aplicar_relleno(
                celda_titulo,
                COLOR_ENCABEZADO
            )

            celda_titulo.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            celda_valor.font = Font(
                bold=True,
                size=18
            )

            celda_valor.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            if etiqueta == "Incidencias críticas":
                color = COLOR_CRITICO

            elif etiqueta in {
                "Advertencias",
                "Puertos sin definir",
                "Switches sin relación"
            }:
                color = COLOR_ADVERTENCIA

            else:
                color = COLOR_INFORMATIVO

            self.aplicar_relleno(
                celda_valor,
                color
            )

            if etiqueta in {
                "Ocupación estimada",
                "Disponibilidad",
                "Cobertura de VLAN"
            }:
                celda_valor.number_format = (
                    "0.00%"
                )

        fila_hallazgos = 25

        hoja.merge_cells(
            start_row=fila_hallazgos,
            start_column=1,
            end_row=fila_hallazgos,
            end_column=6
        )

        celda_hallazgos = hoja.cell(
            row=fila_hallazgos,
            column=1,
            value="PRINCIPALES HALLAZGOS"
        )

        celda_hallazgos.font = Font(
            bold=True,
            color=COLOR_TEXTO_CLARO
        )

        self.aplicar_relleno(
            celda_hallazgos,
            COLOR_TITULO
        )

        hallazgos = (
            self.analizador_reportes
            .obtener_principales_hallazgos()
        )

        fila_actual = (
            fila_hallazgos + 1
        )

        for numero, hallazgo in enumerate(
            hallazgos,
            start=1
        ):
            hoja.merge_cells(
                start_row=fila_actual,
                start_column=1,
                end_row=fila_actual,
                end_column=6
            )

            celda = hoja.cell(
                row=fila_actual,
                column=1,
                value=f"{numero}. {hallazgo}"
            )

            celda.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

            hoja.row_dimensions[
                fila_actual
            ].height = 32

            fila_actual += 1

        fila_actual += 1

        hoja.merge_cells(
            start_row=fila_actual,
            start_column=1,
            end_row=fila_actual,
            end_column=6
        )

        celda_recomendaciones = hoja.cell(
            row=fila_actual,
            column=1,
            value="RECOMENDACIONES"
        )

        celda_recomendaciones.font = Font(
            bold=True,
            color=COLOR_TEXTO_CLARO
        )

        self.aplicar_relleno(
            celda_recomendaciones,
            COLOR_TITULO
        )

        fila_actual += 1

        recomendaciones = (
            self.analizador_reportes
            .obtener_recomendaciones()
        )

        for numero, recomendacion in enumerate(
            recomendaciones,
            start=1
        ):
            hoja.merge_cells(
                start_row=fila_actual,
                start_column=1,
                end_row=fila_actual,
                end_column=6
            )

            celda = hoja.cell(
                row=fila_actual,
                column=1,
                value=(
                    f"{numero}. "
                    f"{recomendacion}"
                )
            )

            celda.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

            self.aplicar_relleno(
                celda,
                COLOR_ADVERTENCIA
            )

            hoja.row_dimensions[
                fila_actual
            ].height = 34

            fila_actual += 1

        for columna in range(
            1,
            7
        ):
            hoja.column_dimensions[
                get_column_letter(
                    columna
                )
            ].width = 18

        hoja.sheet_view.showGridLines = False

    def crear_hoja_dashboard(
        self,
        libro
    ):
        """
        Crea gráficos generales del inventario.
        """
        hoja = libro.create_sheet(
            "Dashboard"
        )

        self.aplicar_titulo(
            hoja,
            "DASHBOARD DEL INVENTARIO DE RED",
            12
        )

        indicadores = (
            self.analizador_reportes
            .obtener_indicadores_generales()
        )

        hoja.cell(
            row=3,
            column=1,
            value="Estado del puerto"
        )

        hoja.cell(
            row=3,
            column=2,
            value="Cantidad"
        )

        self.aplicar_encabezados(
            hoja,
            3,
            [
                "Estado del puerto",
                "Cantidad"
            ]
        )

        datos_ocupacion = [
            (
                "Ocupados",
                indicadores[
                    "puertos_ocupados"
                ]
            ),
            (
                "Disponibles",
                indicadores[
                    "puertos_disponibles"
                ]
            ),
            (
                "Sin definir",
                indicadores[
                    "puertos_sin_definir"
                ]
            )
        ]

        for numero_fila, (
            estado,
            cantidad
        ) in enumerate(
            datos_ocupacion,
            start=4
        ):
            hoja.cell(
                row=numero_fila,
                column=1,
                value=estado
            )

            hoja.cell(
                row=numero_fila,
                column=2,
                value=cantidad
            )

        grafico_ocupacion = DoughnutChart()

        datos = Reference(
            hoja,
            min_col=2,
            min_row=3,
            max_row=6
        )

        categorias = Reference(
            hoja,
            min_col=1,
            min_row=4,
            max_row=6
        )

        grafico_ocupacion.add_data(
            datos,
            titles_from_data=True
        )

        grafico_ocupacion.set_categories(
            categorias
        )

        grafico_ocupacion.title = (
            "Ocupación general de puertos"
        )

        grafico_ocupacion.height = 8
        grafico_ocupacion.width = 13

        hoja.add_chart(
            grafico_ocupacion,
            "D3"
        )

        tipos = (
            self.analizador_reportes
            .obtener_distribucion_tipos()
        )

        fila_tipos = 10

        self.aplicar_encabezados(
            hoja,
            fila_tipos,
            [
                "Tipo de equipo",
                "Cantidad"
            ]
        )

        for indice, item in enumerate(
            tipos,
            start=fila_tipos + 1
        ):
            hoja.cell(
                row=indice,
                column=1,
                value=item["tipo"]
            )

            hoja.cell(
                row=indice,
                column=2,
                value=item["cantidad"]
            )

        if tipos:
            grafico_tipos = BarChart()
            grafico_tipos.type = "bar"
            grafico_tipos.style = 10

            grafico_tipos.title = (
                "Distribución por tipo de equipo"
            )

            grafico_tipos.y_axis.title = (
                "Tipo"
            )

            grafico_tipos.x_axis.title = (
                "Cantidad"
            )

            datos = Reference(
                hoja,
                min_col=2,
                min_row=fila_tipos,
                max_row=(
                    fila_tipos
                    + len(tipos)
                )
            )

            categorias = Reference(
                hoja,
                min_col=1,
                min_row=fila_tipos + 1,
                max_row=(
                    fila_tipos
                    + len(tipos)
                )
            )

            grafico_tipos.add_data(
                datos,
                titles_from_data=True
            )

            grafico_tipos.set_categories(
                categorias
            )

            grafico_tipos.height = 9
            grafico_tipos.width = 16

            hoja.add_chart(
                grafico_tipos,
                "D12"
            )

        vlan = (
            self.analizador_reportes
            .obtener_distribucion_vlan()
        )

        fila_vlan = max(
            fila_tipos + len(tipos) + 3,
            25
        )

        self.aplicar_encabezados(
            hoja,
            fila_vlan,
            [
                "VLAN",
                "Cantidad"
            ]
        )

        for indice, item in enumerate(
            vlan,
            start=fila_vlan + 1
        ):
            hoja.cell(
                row=indice,
                column=1,
                value=item["vlan"]
            )

            hoja.cell(
                row=indice,
                column=2,
                value=item["cantidad"]
            )

        if vlan:
            grafico_vlan = PieChart()

            grafico_vlan.title = (
                "Distribución de VLAN"
            )

            datos = Reference(
                hoja,
                min_col=2,
                min_row=fila_vlan,
                max_row=(
                    fila_vlan
                    + len(vlan)
                )
            )

            categorias = Reference(
                hoja,
                min_col=1,
                min_row=fila_vlan + 1,
                max_row=(
                    fila_vlan
                    + len(vlan)
                )
            )

            grafico_vlan.add_data(
                datos,
                titles_from_data=True
            )

            grafico_vlan.set_categories(
                categorias
            )

            grafico_vlan.height = 9
            grafico_vlan.width = 14

            hoja.add_chart(
                grafico_vlan,
                f"D{fila_vlan}"
            )

        hoja.sheet_view.showGridLines = False
        hoja.column_dimensions["A"].width = 24
        hoja.column_dimensions["B"].width = 14

    def crear_hoja_analisis_sector(
        self,
        libro
    ):
        """
        Crea el análisis de capacidad por sector.
        """
        hoja = libro.create_sheet(
            "Análisis por sector"
        )

        encabezados = [
            "Sector",
            "Puertos",
            "Ocupados",
            "Disponibles",
            "Sin definir",
            "Con VLAN",
            "Sin VLAN",
            "% ocupación",
            "% disponibilidad",
            "% cobertura VLAN",
            "Estado de capacidad"
        ]

        self.aplicar_titulo(
            hoja,
            "CAPACIDAD Y DOCUMENTACIÓN POR SECTOR",
            len(encabezados)
        )

        self.aplicar_encabezados(
            hoja,
            2,
            encabezados
        )

        sectores = (
            self.analizador_reportes
            .obtener_analisis_por_sector()
        )

        for item in sectores:
            disponibilidad = item[
                "disponibilidad_porcentaje"
            ]

            if disponibilidad <= 5:
                estado = "Capacidad crítica"

            elif disponibilidad < 15:
                estado = "Capacidad limitada"

            else:
                estado = "Capacidad disponible"

            hoja.append(
                [
                    item["hoja"],
                    item["total"],
                    item["ocupados"],
                    item["disponibles"],
                    item["sin_definir"],
                    item["con_vlan"],
                    item["sin_vlan"],
                    (
                        item[
                            "ocupacion_porcentaje"
                        ] / 100
                    ),
                    (
                        item[
                            "disponibilidad_porcentaje"
                        ] / 100
                    ),
                    (
                        item[
                            "documentacion_vlan_porcentaje"
                        ] / 100
                    ),
                    estado
                ]
            )

        for fila in range(
            3,
            hoja.max_row + 1
        ):
            for columna in (
                8,
                9,
                10
            ):
                hoja.cell(
                    row=fila,
                    column=columna
                ).number_format = "0.00%"

            estado = hoja.cell(
                row=fila,
                column=11
            ).value

            if estado == "Capacidad crítica":
                color = COLOR_CRITICO

            elif estado == "Capacidad limitada":
                color = COLOR_ADVERTENCIA

            else:
                color = COLOR_CORRECTO

            self.aplicar_relleno(
                hoja.cell(
                    row=fila,
                    column=11
                ),
                color
            )

        self.aplicar_formato_tabla(
            hoja,
            2,
            hoja.max_row,
            len(encabezados)
        )

        self.congelar_y_filtrar(
            hoja,
            2,
            hoja.max_row,
            len(encabezados)
        )

        self.ajustar_columnas(
            hoja
        )

        if sectores:
            grafico = BarChart()
            grafico.type = "col"
            grafico.style = 10

            grafico.title = (
                "Ocupación por sector"
            )

            grafico.y_axis.title = (
                "Cantidad de puertos"
            )

            grafico.x_axis.title = (
                "Sector"
            )

            datos = Reference(
                hoja,
                min_col=3,
                max_col=4,
                min_row=2,
                max_row=(
                    2 + len(sectores)
                )
            )

            categorias = Reference(
                hoja,
                min_col=1,
                min_row=3,
                max_row=(
                    2 + len(sectores)
                )
            )

            grafico.add_data(
                datos,
                titles_from_data=True
            )

            grafico.set_categories(
                categorias
            )

            grafico.height = 11
            grafico.width = 22

            hoja.add_chart(
                grafico,
                "M3"
            )

    def crear_hoja_switches(
        self,
        libro
    ):
        """
        Crea una hoja con información administrativa,
        relaciones y capacidad documentada por switch.
        """
        hoja = libro.create_sheet(
            "Switches"
        )

        encabezados = [
            "IP",
            "Ubicación",
            "Marca",
            "Modelo",
            "MAC",
            "Hoja relacionada",
            "Bloque",
            "Estado relación",
            "Puertos documentados",
            "Ocupados",
            "Disponibles",
            "Sin definir",
            "% ocupación",
            "% disponibilidad",
            "% cobertura VLAN",
            "Estado de capacidad"
        ]

        self.aplicar_titulo(
            hoja,
            "INVENTARIO Y CAPACIDAD DE SWITCHES",
            len(encabezados)
        )

        self.aplicar_encabezados(
            hoja,
            2,
            encabezados
        )

        switches = (
            self.analizador_reportes
            .obtener_analisis_por_switch()
        )

        for switch in switches:
            disponibilidad = switch[
                "disponibilidad_porcentaje"
            ]

            if switch["total_puertos"] == 0:
                estado_capacidad = (
                    "Sin documentación"
                )

            elif disponibilidad <= 5:
                estado_capacidad = (
                    "Capacidad crítica"
                )

            elif disponibilidad < 15:
                estado_capacidad = (
                    "Capacidad limitada"
                )

            else:
                estado_capacidad = (
                    "Capacidad disponible"
                )

            hoja.append(
                [
                    switch["ip"],
                    switch["ubicacion"],
                    switch["marca"],
                    switch["modelo"],
                    switch["mac"],
                    switch["hoja"],
                    switch["bloque"],
                    switch["estado_relacion"],
                    switch["total_puertos"],
                    switch["ocupados"],
                    switch["disponibles"],
                    switch["sin_definir"],
                    (
                        switch[
                            "ocupacion_porcentaje"
                        ] / 100
                    ),
                    (
                        switch[
                            "disponibilidad_porcentaje"
                        ] / 100
                    ),
                    (
                        switch[
                            "cobertura_vlan_porcentaje"
                        ] / 100
                    ),
                    estado_capacidad
                ]
            )

        if not switches:
            hoja.append(
                [
                    "Sin switches registrados"
                ]
            )

        for fila in range(
            3,
            hoja.max_row + 1
        ):
            for columna in (
                13,
                14,
                15
            ):
                hoja.cell(
                    row=fila,
                    column=columna
                ).number_format = "0.00%"

            estado_relacion = hoja.cell(
                row=fila,
                column=8
            ).value

            if estado_relacion == "Relación válida":
                color_relacion = COLOR_CORRECTO

            elif estado_relacion == "Sin relación":
                color_relacion = COLOR_ADVERTENCIA

            else:
                color_relacion = COLOR_CRITICO

            self.aplicar_relleno(
                hoja.cell(
                    row=fila,
                    column=8
                ),
                color_relacion
            )

            estado_capacidad = hoja.cell(
                row=fila,
                column=16
            ).value

            if estado_capacidad == "Capacidad crítica":
                color_capacidad = COLOR_CRITICO

            elif estado_capacidad in {
                "Capacidad limitada",
                "Sin documentación"
            }:
                color_capacidad = COLOR_ADVERTENCIA

            else:
                color_capacidad = COLOR_CORRECTO

            self.aplicar_relleno(
                hoja.cell(
                    row=fila,
                    column=16
                ),
                color_capacidad
            )

        self.aplicar_formato_tabla(
            hoja,
            2,
            hoja.max_row,
            len(encabezados)
        )

        self.congelar_y_filtrar(
            hoja,
            2,
            hoja.max_row,
            len(encabezados)
        )

        self.ajustar_columnas(
            hoja
        )

        if switches:
            grafico = BarChart()
            grafico.type = "col"
            grafico.style = 10

            grafico.title = (
                "Capacidad documentada por switch"
            )

            grafico.y_axis.title = (
                "Cantidad de puertos"
            )

            grafico.x_axis.title = (
                "Switch"
            )

            datos = Reference(
                hoja,
                min_col=10,
                max_col=11,
                min_row=2,
                max_row=(
                    2 + len(switches)
                )
            )

            categorias = Reference(
                hoja,
                min_col=1,
                min_row=3,
                max_row=(
                    2 + len(switches)
                )
            )

            grafico.add_data(
                datos,
                titles_from_data=True
            )

            grafico.set_categories(
                categorias
            )

            grafico.height = 11
            grafico.width = 22

            hoja.add_chart(
                grafico,
                "R3"
            )

    def crear_hoja_validaciones(
        self,
        libro
    ):
        """
        Crea el detalle de incidencias detectadas.
        """
        hoja = libro.create_sheet(
            "Validaciones"
        )

        encabezados = [
            "Nivel",
            "Regla",
            "Hoja",
            "Bloque",
            "Fila Excel",
            "Puerto",
            "Campo",
            "Descripción"
        ]

        self.aplicar_titulo(
            hoja,
            "VALIDACIONES DEL INVENTARIO",
            len(encabezados)
        )

        self.aplicar_encabezados(
            hoja,
            2,
            encabezados
        )

        validaciones = (
            self.validador_inventario
            .ejecutar_validacion()
        )

        for validacion in validaciones:
            hoja.append(
                [
                    validacion.get("nivel"),
                    validacion.get("regla"),
                    validacion.get("hoja"),
                    validacion.get("bloque"),
                    validacion.get("fila_excel"),
                    validacion.get(
                        "puerto_switch"
                    ),
                    validacion.get("campo"),
                    validacion.get(
                        "descripcion"
                    )
                ]
            )

        if not validaciones:
            hoja.append(
                [
                    "CORRECTO",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "No se detectaron incidencias"
                ]
            )

        for fila in range(
            3,
            hoja.max_row + 1
        ):
            nivel = hoja.cell(
                row=fila,
                column=1
            ).value

            if nivel == "CRÍTICO":
                color = COLOR_CRITICO

            elif nivel == "ADVERTENCIA":
                color = COLOR_ADVERTENCIA

            elif nivel == "CORRECTO":
                color = COLOR_CORRECTO

            else:
                color = COLOR_INFORMATIVO

            self.aplicar_color_fila(
                hoja,
                fila,
                len(encabezados),
                color
            )

        self.aplicar_formato_tabla(
            hoja,
            2,
            hoja.max_row,
            len(encabezados)
        )

        self.congelar_y_filtrar(
            hoja,
            2,
            hoja.max_row,
            len(encabezados)
        )

        self.ajustar_columnas(
            hoja,
            ancho_maximo=55
        )

        hoja.column_dimensions["H"].width = 55

    def crear_hoja_inventario(
        self,
        libro,
        inventario
    ):
        """
        Crea el inventario general.
        """
        hoja = libro.create_sheet(
            "Inventario"
        )

        encabezados = [
            "Hoja",
            "Bloque",
            "Fila Excel",
            "Tipo",
            "Equipo",
            "Boca patch",
            "Puerto switch",
            "VLAN"
        ]

        self.aplicar_titulo(
            hoja,
            "INVENTARIO GENERAL",
            len(encabezados)
        )

        self.aplicar_encabezados(
            hoja,
            2,
            encabezados
        )

        for registro in inventario.registros:
            hoja.append(
                [
                    registro.get("hoja"),
                    registro.get("bloque"),
                    registro.get("fila_excel"),
                    registro.get("tipo"),
                    self.valor_visible(
                        registro.get("equipo")
                    ),
                    self.valor_visible(
                        registro.get(
                            "boca_patch"
                        )
                    ),
                    registro.get(
                        "puerto_switch"
                    ),
                    self.valor_visible(
                        registro.get("vlan")
                    )
                ]
            )

        self.aplicar_formato_tabla(
            hoja,
            2,
            hoja.max_row,
            len(encabezados)
        )

        self.congelar_y_filtrar(
            hoja,
            2,
            hoja.max_row,
            len(encabezados)
        )

        self.ajustar_columnas(
            hoja
        )

    def crear_hoja_bloques_incompletos(
        self,
        libro,
        revisor
    ):
        """
        Crea el detalle de campos incompletos.
        """
        hoja = libro.create_sheet(
            "Bloques incompletos"
        )

        encabezados = [
            "Hoja",
            "Bloque",
            "Fila Excel",
            "Puerto switch",
            "Tipo",
            "Equipo",
            "Boca patch",
            "VLAN",
            "Campos faltantes"
        ]

        self.aplicar_titulo(
            hoja,
            "BLOQUES INCOMPLETOS",
            len(encabezados)
        )

        self.aplicar_encabezados(
            hoja,
            2,
            encabezados
        )

        registros = (
            revisor.obtener_filas_incompletas()
        )

        for registro in registros:
            hoja.append(
                [
                    registro.get("hoja"),
                    registro.get("bloque"),
                    registro.get("fila_excel"),
                    registro.get(
                        "puerto_switch"
                    ),
                    self.valor_visible(
                        registro.get("tipo")
                    ),
                    self.valor_visible(
                        registro.get("equipo")
                    ),
                    self.valor_visible(
                        registro.get(
                            "boca_patch"
                        )
                    ),
                    self.valor_visible(
                        registro.get("vlan")
                    ),
                    ", ".join(
                        registro.get(
                            "campos_faltantes",
                            []
                        )
                    )
                ]
            )

        if not registros:
            hoja.append(
                [
                    "Sin pendientes",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "No se encontraron campos incompletos"
                ]
            )

        self.aplicar_formato_tabla(
            hoja,
            2,
            hoja.max_row,
            len(encabezados)
        )

        self.congelar_y_filtrar(
            hoja,
            2,
            hoja.max_row,
            len(encabezados)
        )

        self.ajustar_columnas(
            hoja
        )

        hoja.column_dimensions["I"].width = 42

    def crear_hoja_duplicados(
        self,
        libro,
        inventario
    ):
        """
        Crea el detalle de puertos repetidos.
        """
        hoja = libro.create_sheet(
            "Puertos repetidos"
        )

        encabezados = [
            "Hoja",
            "Bloque",
            "Puerto",
            "Primera fila",
            "Segunda fila",
            "Primer equipo",
            "Segundo equipo"
        ]

        self.aplicar_titulo(
            hoja,
            "PUERTOS REPETIDOS PARA REVISIÓN",
            len(encabezados)
        )

        self.aplicar_encabezados(
            hoja,
            2,
            encabezados
        )

        duplicados = (
            inventario.duplicados_detectados
        )

        for duplicado in duplicados:
            hoja.append(
                [
                    duplicado.get("hoja"),
                    duplicado.get("bloque"),
                    duplicado.get(
                        "puerto_switch"
                    ),
                    duplicado.get(
                        "primera_fila"
                    ),
                    duplicado.get(
                        "segunda_fila"
                    ),
                    duplicado.get(
                        "primer_equipo"
                    ),
                    duplicado.get(
                        "segundo_equipo"
                    )
                ]
            )

        if not duplicados:
            hoja.append(
                [
                    "Sin duplicados",
                    "",
                    "",
                    "",
                    "",
                    "",
                    ""
                ]
            )

        self.aplicar_formato_tabla(
            hoja,
            2,
            hoja.max_row,
            len(encabezados)
        )

        self.congelar_y_filtrar(
            hoja,
            2,
            hoja.max_row,
            len(encabezados)
        )

        if duplicados:
            for fila in range(
                3,
                hoja.max_row + 1
            ):
                self.aplicar_color_fila(
                    hoja,
                    fila,
                    len(encabezados),
                    COLOR_CRITICO
                )

        self.ajustar_columnas(
            hoja
        )

    def exportar_reporte(
        self,
        inventario
    ):
        """
        Genera el informe completo de administración
        de red.
        """
        if not inventario.registros:
            print(
                "\nNo existen registros para generar "
                "el reporte."
            )
            return None

        if self.analizador_reportes is None:
            print(
                "\nNo se configuró el analizador "
                "de reportes."
            )
            return None

        if self.validador_inventario is None:
            print(
                "\nNo se configuró el validador "
                "del inventario."
            )
            return None

        revisor = RevisorIncompletos(
            inventario
        )

        libro = Workbook()
        ruta_salida = None

        try:
            self.crear_hoja_resumen_ejecutivo(
                libro
            )

            self.crear_hoja_dashboard(
                libro
            )

            self.crear_hoja_analisis_sector(
                libro
            )

            self.crear_hoja_switches(
                libro
            )

            self.crear_hoja_validaciones(
                libro
            )

            self.crear_hoja_inventario(
                libro,
                inventario
            )

            self.crear_hoja_bloques_incompletos(
                libro,
                revisor
            )

            self.crear_hoja_duplicados(
                libro,
                inventario
            )

            marca_tiempo = datetime.now().strftime(
                "%Y%m%d_%H%M%S"
            )

            nombre_archivo = (
                f"reporte_"
                f"{NOMBRE_APLICACION.lower()}_"
                f"{marca_tiempo}.xlsx"
            )

            ruta_salida = (
                self.carpeta_reportes
                / nombre_archivo
            )

            libro.save(
                ruta_salida
            )

        except PermissionError:
            print(
                "\nNo se pudo guardar el reporte. "
                "Comprueba que no esté abierto en Excel."
            )
            return None

        except Exception as error:
            print(
                "\nNo se pudo generar el reporte."
            )

            print(
                f"Detalle: {error}"
            )

            return None

        finally:
            libro.close()

        print(
            "\nReporte generado correctamente:"
        )

        print(
            ruta_salida
        )

        return ruta_salida